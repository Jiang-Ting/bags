# -*- coding: utf-8 -*-
import numpy as np
import matplotlib.pyplot as plt
import time
import pandas as pd
import openpyxl
import queue
import math

# 动态规划法
def bag(n, c, w, v):
    # 置零，表示初始状态
    value = [[0 for j in range(c + 1)] for i in range(n + 1)]
    for i in range(1, n + 1):
        for j in range(1, c + 1):
            value[i][j] = value[i - 1][j]
            # 背包总容量够放当前物体，遍历前一个状态考虑是否置换
            if j >= w[i - 1] and value[i][j] < value[i - 1][j - w[i - 1]] + v[i - 1]:
                value[i][j] = value[i - 1][j - w[i - 1]] + v[i - 1]
    # for x in value:
    #     print(x)
    return value

def show(n, c, w, value):
    print('最大价值为:', value[n][c])
    x = [False for i in range(n)]
    j = c
    for i in range(n, 0, -1):
        if value[i][j] > value[i - 1][j]:
            x[i - 1] = True
            j -= w[i - 1]
    print('背包中所装物品为:')
    for i in range(n):
        print(i+1, ' ', end='')
    print()
    for i in range(n):
        if x[i]:
            print('1', ' ', end='')
        else:
            print('0', ' ', end='')
        # if x[i]:
        #     print('第', i+1, '个,', end='')
    f = open('out.txt', 'w', encoding='utf-8')
    for i in x:
        f.write(str(i) + str(x[i])+ '\n')
    f.close()
    data = pd.read_csv("out.txt", sep="\t")
    wb = openpyxl.load_workbook(r'out.xlsx')
    ws = wb['Sheet1']
    # 取出distance_list列表中的每一个元素，openpyxl的行列号是从1开始取得，所以我这里i从1开始取
    ws.cell(row=1, column=2).value ='物品'
    ws.cell(row=1, column=3).value ='是否装入'
    for i in range(1, len(x) + 1):
        distance = x[i - 1]
        # 写入位置的行列号可以任意改变，这里我是从第2行开始按行依次插入第11列
        ws.cell(row=i + 1, column=3).value = distance
        distance = i
        # 写入位置的行列号可以任意改变，这里我是从第2行开始按行依次插入第11列
        ws.cell(row=i + 1, column=2).value = distance
    # 保存操作
    wb.save(r'out1.xlsx')


# 快排的主函数，传入参数为一个列表，左右两端的下标
def QuickSort(list, v, w, low, high):
    if high > low:
        # 传入参数，通过Partitions函数，获取k下标值
        k = Partitions(list, v, w, low, high)
        # 递归排序列表k下标左侧的列表
        QuickSort(list, v, w, low, k - 1)
        # 递归排序列表k下标右侧的列表
        QuickSort(list, v, w, k + 1, high)


def Partitions(r, v, w, low, high):
    i = low
    j = high
    # 当left下标，小于right下标的情况下，此时判断二者移动是否相交，若未相交，则一直循环
    while i < j:
        while i < j and r[i] >= r[j]:
            j -= 1
        if i < j:
            temp = r[i]
            r[i] = r[j]
            r[j] = temp

            temp = v[i]
            v[i] = v[j]
            v[j] = temp

            temp = w[i]
            w[i] = w[j]
            w[j] = temp

            i += 1
        while i < j and r[i] >= r[j]:
            i += 1
        if i < j:
            temp = r[i]
            r[i] = r[j]
            r[j] = temp

            temp = v[i]
            v[i] = v[j]
            v[j] = temp

            temp = w[i]
            w[i] = w[j]
            w[j] = temp
        j -= 1
    return i


# 贪心法
def KnapSack(w, v, n, c):
    x = [0 for j in range(n+1)]
    maxValue = 0
    i = 0
    for i in range(0, n):
        if w[i] < c:
            x[i] = 1
            maxValue += v[i]
            c = c - w[i]
            # print(i)
        else:
            break
    # print('=========')
    # print(i)
    x[i] = float(c/w[i])
    maxValue += x[i]*v[i]
    return maxValue


# 回溯法
def test(capacity, w, v):
    vec_len = 2 ** (len(v) + 1) - 1  # tree `s size
    # vec_len = 10000000
    vec_v = [0 for j in range(vec_len)]
    vec_w = [0 for j in range(vec_len)]
    # print(vec_v)
    # vec_v = np.zeros(vec_len)
    # vec_w = np.zeros(vec_len)
    vec_w[0] = capacity
    que = queue.Queue()
    que.put(0)
    best = 0
    while (not que.empty()):
        current = que.get()
        level = int(math.log(current + 1, 2))
        if (vec_v[current] > vec_v[best]):
            best = current

        left = 2 * current + 1  # left child   index
        right = 2 * current + 2  # right child index

        if (left < vec_len and vec_w[current] - w[level] > 0 and vec_v[current] + v[level] > vec_v[best]):
            vec_v[left] = int(vec_v[current] + v[level])
            vec_w[left] = vec_w[current] - w[level]
            que.put(left)
        if (right < vec_len and sum(v[level + 1:-1]) + vec_v[current] > vec_v[best]):
            vec_v[right] = vec_v[current]
            vec_w[right] = vec_w[current]
            que.put(right)
    # print(vec_w[best], vec_v[best])
    print(vec_v[best])

# bestV = 0
# curW = 0
# curV = 0
# bestx = None
#
#
# def backtrack(i, n, w, v, c):
#     global bestV, curW, curV, x, bestx
#     if i >= n:
#         if bestV < curV:
#             bestV = curV
#             bestx = x[:]
#     else:
#         if curW + w[i] <= c:
#             x[i] = True
#             curW += w[i]
#             curV += v[i]
#             backtrack(i + 1, n, w, v, c)
#             curW -= w[i]
#             curV -= v[i]
#         x[i] = False
#         backtrack(i + 1, n, w, v, c)



if __name__ == "__main__":
    # 读取txt文件,按行读取后删除'\n'
    # 方法一
    # with open('test.txt','r') as file:
    #     content_list = file.readlines()
    #     contentall = [x.strip() for x in content_list]
    #     print(contentall)

    # 方法二
    # with open('data\beibao0.in', 'r') as f1:
    #     list1 = f1.readlines()
    # for i in range(0, len(list1))
    #     list1[i] = list1[i].rstrip('\n')

    # 方法三
    file = 'data\\beibao9.in'
    option = int(input('数据集(0-9):'))
    if option == 0:
        file = 'data\\beibao0.in'
    if option == 1:
        file = 'data\\beibao1.in'
    if option == 2:
        file = 'data\\beibao2.in'
    if option == 3:
        file = 'data\\beibao3.in'
    if option == 4:
        file = 'data\\beibao4.in'
    if option == 5:
        file = 'data\\beibao5.in'
    if option == 6:
        file = 'data\\beibao6.in'
    if option == 7:
        file = 'data\\beibao7.in'
    if option == 8:
        file = 'data\\beibao8.in'
    if option == 9:
        file = 'data\\beibao9.in'
    # print(file)
    f = open(file, 'r', encoding='utf-8')
    arr = []
    for m1 in f:
        m2 = m1.strip("\n")
        arr.append(m2)
    del f, m1
    arr_weight = []
    arr_value = []
    for row in arr:
        arr_weight.append(row.split()[0])
        arr_value.append(row.split()[1])  # 用空格将列表项拆分开
    arr_value = list(map(int, arr_value))  # 列表项为字符串，强制转换为整数型
    arr_weight = list(map(int, arr_weight))
    nums = arr_value[0]
    capacity = arr_weight[0]

    # 删除第一行物品个数和背包容量，列表整体前移
    for i in range(len(arr_value)-1):
        arr_value[i] = arr_value[i+1]
        arr_weight[i] = arr_weight[i+1]
    del arr_value[nums-1]
    del arr_weight[nums-1]
    # print(arr_value)
    # print(arr_weight)

    # 画散点图
    plt.rcParams['font.sans-serif'] = ['SimHei']
    plt.rcParams['axes.unicode_minus'] = False
    # matplotlib画图中中文显示会有问题，需要这两行设置默认字体
    # rcParams用来设置画图时的一些基本参数

    # 横轴与纵轴名称及大小
    plt.xlabel('Weight')
    plt.ylabel('Value')
    plt.xlim(xmax=110, xmin=0)
    plt.ylim(ymax=110, ymin=0)

    # 点的颜色
    colors2 = '#DC143C'

    # 点面积
    area = np.pi * 2 ** 2

    plt.scatter(arr_weight, arr_value, s=area, c=colors2, alpha=0.4, label='物品')
    plt.legend()  # 显示字符或表达式向量
    plt.savefig(r'D:\Desktop\12345svm.png', dpi=300)
    plt.show()

    # 计算性价比
    ratio = []
    for i in range(len(arr_value)):
        b = round((arr_value[i]/arr_weight[i]), 4)
        ratio.append(b)
        # print('价值：' + str(arr_value[i]) + ' ' + '重量：' + str(arr_weight[i]) + ' ' + '性价比：' + str(b))

    # 利用快速排序算法，按性价比非递增排序
    QuickSort(ratio, arr_value, arr_weight, 0, nums-1)
    for i in range(len(arr_value)):
        print('价值：' + str(arr_value[i]) + ' ' + '重量：' + str(arr_weight[i]) + ' ' + '性价比：' + str(ratio[i]))

    option = int(input('1贪心法 2动态规划法 3回溯法'+'\n'))
    # option = 3
    start = time.perf_counter()
    # while option:
    # 贪心法
    if option == 1:
        value = KnapSack(arr_weight, arr_value, nums, capacity)
        value = round(value, 4)
        print(value)

    # 动态规划法
    if option == 2:
        value = bag(nums, capacity, arr_weight, arr_value)
        show(nums, capacity, arr_weight, value)
    # print()

    # 回溯法
    if option == 3:
        test(capacity, arr_weight, arr_value)
        print()
        # bestV = 0
        # curW = 0
        # curV = 0
        # bestx = None
        # x = [False for i in range(nums)]
        # backtrack(0, nums, arr_weight, arr_value, capacity)
        # print(bestV)
        # print(bestx)
        # option = int(input('1贪心法 2动态规划法 3回溯法' + '\n'))

    end = time.perf_counter()
    print()
    print('running time: %.8s s' % ((end - start)))

    # 结果写入文件
    f = open('out.txt', 'w', encoding='utf-8')
    f.write('性价比')
    f.write('\n')
    for i in ratio:
        f.write(str(i) + '\n')
    f.write(str(end-start)+ 's'+ '\n')
    f.close()
    data = pd.read_csv("out.txt", sep="\t")
    data.to_excel("out.xlsx", index=False)




